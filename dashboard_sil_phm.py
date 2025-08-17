import tkinter as tk
from tkinter import filedialog, messagebox, ttk, simpledialog
from tkcalendar import DateEntry
import pandas as pd
from openpyxl import load_workbook
import os
import ast
from datetime import datetime
import shutil
import customtkinter
import sys
import numpy as np
from openpyxl import Workbook

# Di bagian atas script, setelah import modules
calculated_and_printed = {}  # Dictionary untuk melacak status perhitungan dan pencetakan

printed = False  # Status untuk mengecek apakah data telah di-print

# Global variable to store the filepath of the loaded Excel file
loaded_file_path = None

user_name = None

def show_initial_dialog():
    global user_name
    last_update_info = "None"
    last_user_info = "None"

    # Cek apakah ada informasi update terakhir
    log_excel_path = 'log_data_sil_phm.xlsx'
    log_sheet_name = 'Log Update Database'
    if os.path.exists(log_excel_path):
        workbook = load_workbook(log_excel_path)
        if log_sheet_name in workbook.sheetnames:
            log_sheet = workbook[log_sheet_name]
            if log_sheet.max_row > 1:
                last_row = log_sheet.max_row
                last_update_info = log_sheet.cell(row=last_row, column=6).value or "None"
                last_user_info = log_sheet.cell(row=last_row, column=5).value or "None"
    
    while True:  # Tambahkan loop
        user_input = customtkinter.CTkInputDialog(
            title="Welcome to PHM SIL Dashboard Calculation",
            text=f"Last Update Database was on: {last_update_info}\n"
                 f"Last User: {last_user_info}\n"
                 "Please enter your name:"
        )
        user_name = user_input.get_input()  # Asumsi metode ini ada
        if user_name is None:  # Dialog telah ditutup
            sys.exit()  # Keluar dari aplikasi
        elif user_name.strip():  # Nama valid
            break  # Keluar dari loop
        else:  # Nama kosong
            messagebox.showwarning("Warning", "Name cannot be empty. Please enter your name.")  # Tampilkan peringatan

# Function to load data from the uploaded Excel file
def load_uploaded_data(filepath):
    global loaded_file_path
    try:
        # When reading the Excel file, ensure the 'Package' column is read as strings
        df = pd.read_excel(filepath, dtype={'Package': str})

        # Define the expected types for each column
        expected_types = {
            'Platform': str,
            'Package': str,
            'Critical': str,
            'Number of Devices': int,
            'Operating Years': int,
            'Sum of Test': int,
            'On Demand Fail': int,
            'Last Test': 'datetime64[ns]',
        }

        # Check if all required columns exist
        missing_columns = [col for col in expected_types if col not in df.columns]
        if missing_columns:
            messagebox.showerror(
                "Error",
                f"The uploaded file is missing columns: {', '.join(missing_columns)}. Please upload the correct database."
            )
            return pd.DataFrame()

        # Verify the type of each column
        for col, expected_type in expected_types.items():
            if expected_type == int:
                if not all(isinstance(x, (int, float, np.int64, np.float64)) and not np.isnan(x) for x in df[col]):
                    messagebox.showerror("Error", f"Column '{col}' should only contain integer values.")
                    return pd.DataFrame()
            elif expected_type == str:
                if not all(isinstance(x, str) for x in df[col]):
                    messagebox.showerror("Error", f"Column '{col}' should only contain string values.")
                    return pd.DataFrame()
            elif expected_type == 'datetime64[ns]':
                if not pd.api.types.is_datetime64_ns_dtype(df[col]):
                    messagebox.showerror("Error", f"Column '{col}' should only contain date values.")
                    return pd.DataFrame()

        # Additional check for 'Test on' columns
        test_on_columns = [col for col in df.columns if col.startswith('Test on')]
        if any(not pd.api.types.is_integer_dtype(df[col]) for col in test_on_columns):
            bad_columns = [col for col in test_on_columns if not pd.api.types.is_integer_dtype(df[col])]
            messagebox.showerror("Error", f"Columns '{', '.join(bad_columns)}' should only contain integer values.")
            return pd.DataFrame()
        
        platform_choices = df['Platform'].unique()
        
        # Process the 'Package' column to merge 'CRITICAL' and 'SDV' if they are separate
        package_choices = df['Package'].unique()
        corrected_packages = []
        for package in package_choices:
            # This is where you can handle the merging logic, if necessary
            if package.strip() in ["CRITICAL", "SDV"] and "CRITICAL SDV" not in corrected_packages:
                # If CRITICAL or SDV is found and CRITICAL SDV is not already added, add it
                corrected_packages.append("CRITICAL SDV")
            else:
                corrected_packages.append(package)
        
        # Update the dropdowns with the corrected list
        update_dropdowns(platform_choices, corrected_packages)
        loaded_file_path = filepath  # Store the path to the loaded file
        return df
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
        return pd.DataFrame()

# Update the dropdown menus with data from the uploaded file
def update_dropdowns(platform_choices, package_choices):
    # Ensure platform_choices and package_choices are proper lists
    try:
        platform_choices = [ast.literal_eval(item) if isinstance(item, str) and item.startswith("[") else item for item in platform_choices]
        package_choices = [ast.literal_eval(item) if isinstance(item, str) and item.startswith("[") else item for item in package_choices]
    except ValueError:
        # If there's an error with literal_eval, pass and use the value as is
        pass
    
    # Flatten the lists if they contain sublists
    platform_choices = [item for sublist in platform_choices for item in (sublist if isinstance(sublist, list) else [sublist])]
    package_choices = [item for sublist in package_choices for item in (sublist if isinstance(sublist, list) else [sublist])]

    # Update combobox values
    platform_menu.configure(values=platform_choices)
    package_menu.configure(values=package_choices)
    sil_platform_menu.configure(values=platform_choices)

    # Optionally, clear any previous selection
    platform_menu.set('')
    package_menu.set('')
    sil_platform_menu.set('')

# Function to pack the remaining widgets into the sidebar
def pack_remaining_widgets():
    # Pack all widgets that should be visible after the file is uploaded
    upload_success_label.pack(fill='x', pady=(5, 0)) 
    update_label.pack(fill='x', pady=5)
    platform_label.pack(fill='x', pady=5)
    platform_menu.pack(fill='x', pady=5)
    package_label.pack(fill='x', pady=5)
    package_menu.pack(fill='x', pady=5)
    last_test_label.pack(fill='x', pady=5)
    last_test_entry.pack(fill='x', pady=5)
    pass_button.pack(fill='x', pady=5)
    fail_button.pack(fill='x', pady=5)
    sil_calculation_label.pack(fill="x", pady=5)
    sil_platform_label.pack(fill='x', pady=5)
    sil_platform_menu.pack(fill='x', pady=5)

# Function to handle file uploads and to pack other widgets after successful upload
def upload_file():
    # Clear the previous widgets
    for widget in sidebar.winfo_children():
        widget.pack_forget()
    
    upload_label.pack(fill='x')
    # Show only the upload button
    upload_button.pack(fill='x', pady=5)
    
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        df = load_uploaded_data(file_path)
        if df.empty:
            return
        
        # Once the file is loaded, pack the rest of the widgets
        pack_remaining_widgets()
        
        # Tampilkan pesan sukses di bawah tombol "Browse Files"
        file_name = os.path.basename(file_path)
        upload_success_label.configure(text=f" '{file_name}' berhasil diupload!")

# Function to update Excel file based on the selections
def update_excel_file(pass_or_fail):
    global loaded_file_path
    if not loaded_file_path:
        messagebox.showerror("Error", "Please upload a database file first.")
        return

    platform = platform_var.get()
    package = package_var.get()
    # Retrieve the date from the DateEntry widget
    last_test_date = last_test_entry.get_date()

    # Check if the date is before the allowed minimum date
    min_allowed_date = datetime(2022, 12, 31).date()
    if last_test_date < min_allowed_date:
        messagebox.showerror("Error", "Last test date cannot be before December 31, 2022.")
        return
    
    # Check if the platform and package are selected
    if not platform or not package:
        messagebox.showerror("Error", "Please select a platform and a package.")
        return

    # Check if the file has been uploaded
    if not loaded_file_path:
        messagebox.showerror("Error", "Please upload an Excel file first.")
        return

    try:
        # Check if the Excel file is open
        if os.path.isfile(loaded_file_path):
            try:
                os.rename(loaded_file_path, loaded_file_path)  # Can we rename the file (check lock)?
            except OSError as e:
                messagebox.showerror("Error", "Close the Excel file before updating.")
                return
        df = pd.read_excel(loaded_file_path)
        # Pastikan 'Last Test' dan 'Start Time' adalah kolom tanggal
        df['Last Test'] = pd.to_datetime(df['Last Test']).dt.date
        df['Start Time'] = pd.to_datetime(df['Start Time']).dt.date  # Konversi 'Start Time' ke tanggal saja

        # Mencari baris dengan Platform dan Package yang dipilih
        row_indices = df[(df['Platform'] == platform) & (df['Package'] == package)].index

        # Memperbarui Last Test dan menghitung Operating Years
        start_time = df.loc[row_indices, 'Start Time'].min()  # Ambil tanggal terawal sebagai start time
        # Use the correct variable for the date when updating the DataFrame
        df.loc[row_indices, 'Last Test'] = last_test_date  # Use last_test_date here
        df.loc[row_indices, 'Operating Years'] = (last_test_date - start_time).days // 365

        # Mengecek dan membuat kolom baru untuk tahun dari last_test_date di posisi yang benar
        year = last_test_date.year
        column_name = f'Test on {year}'
        
        # Dapatkan semua kolom 'Test on' dan urutkan berdasarkan tahun
        test_on_columns = [col for col in df.columns if col.startswith('Test on ')]
        test_on_years = sorted([int(col.split(' ')[2]) for col in test_on_columns])
        
        if column_name not in df.columns:
            # Tentukan posisi kolom berdasarkan tahun
            insert_position = 0  # Default ke posisi awal jika belum ada kolom 'Test on'
            for test_year in test_on_years:
                if year > test_year:
                    # Dapatkan posisi kolom tahun ini untuk menentukan posisi sisipan
                    insert_position = df.columns.get_loc(f'Test on {test_year}') + 1
            
            # Sisipkan kolom baru di posisi yang benar
            df.insert(insert_position, column_name, 0)
        
        df.loc[row_indices, column_name] += 1  # Hanya update kolom tahun yang sesuai dengan last_test_date

        # Hitung ulang 'Sum of Test' setelah semua update
        test_columns = [col for col in df.columns if col.startswith('Test on ')]
        df.loc[row_indices, 'Sum of Test'] = df.loc[row_indices, test_columns].sum(axis=1)

        # Determine the value for the 'fail' column based on the button clicked
        fail = pass_or_fail == "fail"
        fail_input_value = fail_input_var.get() if fail else 0

        # Cek jika tombol Fail diklik
        if pass_or_fail == "fail":
            # Pastikan 'On Demand Fail' ada, jika tidak, inisialisasi dengan 0
            if 'On Demand Fail' not in df.columns:
                df['On Demand Fail'] = 0
            # Menambahkan angka sesuai fail_input ke 'On Demand Fail' untuk baris yang dipilih
            df.loc[row_indices, 'On Demand Fail'] = fail_input_value
        # Save the updated DataFrame back to the Excel file
        df.to_excel(loaded_file_path, index=False)
        hide_on_demand_fail_input()
        # Tampilkan notifikasi sukses dengan timestamp
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        messagebox.showinfo("Success", f"Database updated with a {pass_or_fail} result at {current_time}.")

        # Tambahkan log ke sheet "Log Update Database"
        add_log_update_database(platform, package, last_test_date, pass_or_fail)
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while updating the Excel file: {e}")
        print(f"Exception: {e}")  # Print the exception to the console for debugging

# Pass and Fail buttons with logic
def pass_test():
    platform = platform_var.get()
    package = package_var.get()
    last_test_date = last_test_entry.get_date().strftime('%Y-%m-%d')
    # Konfirmasi dari pengguna dengan detail yang lebih spesifik
    confirm = messagebox.askyesno(
        "Konfirmasi",
        f"Apakah Anda yakin ingin mengupdate Platform {platform}, Package {package}, "
        f"dan Last Test Date pada {last_test_date} sebagai PASS?"
    )
    if confirm:
        update_excel_file("pass")

# Modify the fail_test function to show the On Demand Fail input
def fail_test():
    # Memperlihatkan input dan tombol konfirmasi
    show_on_demand_fail_input()

# Function to show the On Demand Fail input
def show_on_demand_fail_input():
    # Get the selected package and platform
    platform = platform_var.get()
    package = package_var.get()

    if not platform or not package:
        messagebox.showerror("Error", "Please select a platform and a package first.")
        fail_button.configure(state='normal')  # Re-enable the 'Fail' button here
        return

    # Retrieve the 'Number of Devices' for the selected package and platform
    df = pd.read_excel(loaded_file_path)  # Make sure to handle potential errors here
    num_devices = df.loc[(df['Platform'] == platform) & (df['Package'] == package), 'Number of Devices'].values[0]
    
    # Update the maximum value for the On Demand Fail input
    fail_input.configure(to=num_devices)
    
    # Show the On Demand Fail input and its label
    fail_input_label.pack(fill='x', pady=5)
    fail_input.pack(fill='x', pady=5)

    # Only show the confirm button after the fail input is shown
    confirm_fail_button.pack(fill='x', pady=5)

    # Paksa pengguna untuk memasukkan input sebelum mengonfirmasi
    confirm_fail_button.configure(state='normal')

    # Memperlihatkan input dan tombol konfirmasi tepat di bawah tombol "Fail"
    fail_input_label.pack(fill='x', pady=5, after=fail_button)
    fail_input.pack(fill='x', pady=5, after=fail_input_label)
    confirm_fail_button.pack(fill='x', pady=5, after=fail_input)

# Fungsi untuk menyembunyikan input On Demand Fail
def hide_on_demand_fail_input():
    # Sembunyikan input dan labelnya
    fail_input_label.pack_forget()
    fail_input.pack_forget()
    # Sembunyikan tombol konfirmasi
    confirm_fail_button.pack_forget()
    # Aktifkan kembali tombol Fail
    fail_button.configure(state='normal')

# This function will be called when the "Confirm Fail Update" button is clicked
def confirm_fail_update():
    platform = platform_var.get()
    package = package_var.get()
    last_test_date = last_test_entry.get_date().strftime('%Y-%m-%d')
    on_demand_fail = fail_input_var.get()
    # Konfirmasi dari pengguna dengan detail yang lebih spesifik
    confirm = messagebox.askyesno(
        "Konfirmasi",
        f"Apakah Anda yakin ingin mengupdate Platform {platform}, Package {package}, "
        f"dan Last Test Date pada {last_test_date} itu memiliki On Demand Fail sebanyak {on_demand_fail}?"
    )
    if confirm:
        try:
            update_excel_file("fail")
        except Exception as e:
            # If update fails, re-enable the 'Fail' button
            fail_button.configure(state='normal')
            messagebox.showerror("Error", f"An error occurred while updating the Excel file: {e}")
        else:
            # If update succeeds, hide the input and confirmation button
            hide_on_demand_fail_input() 

results_treeview = None
scrollbar = None
df_results = pd.DataFrame(columns=[
    "Timestamp", "Sites", "Equipment", "Current Package PM Test Interval", 
    "Instrument Type", "Current Instr PM Test Interval", 
    "Current Probability Failure on Demand Average", 
    "SIL Class - Current ESD1 Test Interval (Baseline 1Y)", 
    "Forecast SIL Class at 2Y ESD1 Test Interval", 
    "Compliance of SIL at forecast 2Y ESD1 test interval", 
    "Proposed Package", "Reliability", "Availability"
])

def add_calculation_result(timestamp, sites, equipment, current_package_pm_test_interval,
                           instrument_type, current_instr_pm_test_interval,
                           current_pfod, sil_class, forecast_sil_class, compliance, proposed_package,
                           reliability, availability):
    global df_results
    # Membuat dictionary untuk baris baru
    new_row = {
        "Timestamp": timestamp,
        "Sites": sites,
        "Equipment": f"ESD-1 {sites}",
        "Current Package PM Test Interval": current_package_pm_test_interval,
        "Instrument Type": instrument_type,
        "Current Instr PM Test Interval": current_instr_pm_test_interval,
        "Current Probability Failure on Demand Average": current_pfod,
        "SIL Class - Current ESD1 Test Interval (Baseline 1Y)": sil_class,
        "Forecast SIL Class at 2Y ESD1 Test Interval": forecast_sil_class,
        "Compliance of SIL at forecast 2Y ESD1 test interval": compliance,
        "Proposed Package": proposed_package,
        "Reliability": f"{reliability}%",
        "Availability": f"{availability}%"
    }
    # Membuat DataFrame sementara dari dictionary
    temp_df = pd.DataFrame([new_row])
    # Menggabungkan DataFrame sementara dengan df_results
    df_results = pd.concat([df_results, temp_df], ignore_index=True)

def add_or_update_calculation_result(timestamp, sites, equipment, current_package_pm_test_interval,
                                     instrument_type, current_instr_pm_test_interval,
                                     current_pfod, sil_class, forecast_sil_class, compliance, 
                                     proposed_package, reliability, availability):
    global df_results
    # Cek apakah hasil untuk 'sites' dan 'equipment' sudah ada
    existing_rows = df_results[(df_results['Sites'] == sites) & (df_results['Equipment'] == equipment)]
    if not existing_rows.empty:
        # Jika ada, update baris tersebut dengan data baru
        idx = existing_rows.index
        df_results.loc[idx, 'Timestamp'] = timestamp
        df_results.loc[idx, 'Current Package PM Test Interval'] = current_package_pm_test_interval
        df_results.loc[idx, 'Instrument Type'] = instrument_type
        df_results.loc[idx, 'Current Instr PM Test Interval'] = current_instr_pm_test_interval
        df_results.loc[idx, 'Current Probability Failure on Demand Average'] = current_pfod
        df_results.loc[idx, 'SIL Class - Current ESD1 Test Interval (Baseline 1Y)'] = sil_class
        df_results.loc[idx, 'Forecast SIL Class at 2Y ESD1 Test Interval'] = forecast_sil_class
        df_results.loc[idx, 'Compliance of SIL at forecast 2Y ESD1 test interval'] = compliance
        df_results.loc[idx, 'Proposed Package'] = proposed_package
        df_results.loc[idx, 'Reliability'] = reliability
        df_results.loc[idx, 'Availability'] = availability
    else:
        # Jika tidak ada, tambahkan baris baru
        add_calculation_result(timestamp, sites, equipment, current_package_pm_test_interval,
                               instrument_type, current_instr_pm_test_interval,
                               current_pfod, sil_class, forecast_sil_class, compliance, 
                               proposed_package, reliability, availability)

main_content = None

def build_results_treeview(parent):
    # Create a frame for the Treeview and the scrollbars
    tree_frame = customtkinter.CTkFrame(parent)
    tree_frame.pack(fill="both", expand=True, pady=10)

    # Create the Treeview widget
    tree = ttk.Treeview(tree_frame, columns=list(df_results.columns), show='headings')
    tree.pack(side="left", fill="both", expand=True)

    # Configure style for Treeview
    style = ttk.Style()
    style.configure("Treeview", font=('Arial', 15))  # Change 'Helvetica' and '10' to your desired font and size
    style.configure("Treeview.Heading", font=('Arial', 15, 'bold'))  # Change font for column headings

    # Configure the column headings
    for col in df_results.columns:
        tree.heading(col, text=col)
        # Set the column's width and anchor
        tree.column(col, width=50)  # Adjust width as needed

    return tree

def autoresize_columns(treeview):
    for column in treeview["columns"]:
        treeview.column(column, width=customtkinter.CTkFont().measure(column.title()))
        max_length = max([customtkinter.CTkFont().measure(str(value)) for value in df_results[column].astype(str)])
        treeview.column(column, width=max_length)

# Function to create and update the Treeview with calculation results
def update_results_treeview():
    if results_treeview is not None:
        results_treeview.delete(*results_treeview.get_children())
        for index, row in df_results.iterrows():
            results_treeview.insert('', 'end', values=tuple(row))
            for i, val in enumerate(row):
                col_width = customtkinter.CTkFont().measure(str(val))
                if results_treeview.column(df_results.columns[i], width=None) < col_width:
                    results_treeview.column(df_results.columns[i], width=col_width)

# Fungsi untuk membangun dan memperbarui Treeview dengan hasil perhitungan
def build_and_update_results_treeview():
    global results_treeview
    # Check if the Treeview exists before trying to delete its children or create a new one
    if results_treeview is not None and results_treeview.winfo_exists():
        results_treeview.delete(*results_treeview.get_children())
        update_results_treeview()
        autoresize_columns(results_treeview)
    else:
        # Create the Treeview again because the old one was destroyed
        results_treeview = build_results_treeview(main_content)
        update_results_treeview()
        autoresize_columns(results_treeview)

def calculate_failure_rates(fail, operating_time_hours):
    if operating_time_hours == 0:
        return 0
    return (fail / operating_time_hours) * 1_000_000

def calculate_pfd(on_demand_fail, operating_time_years, sum_of_test, number_of_devices):
    if sum_of_test == 0:  # Hindari pembagian dengan nol
        return 0
    operating_time_hours = operating_time_years * 8760 * number_of_devices
    failure_rates = calculate_failure_rates(on_demand_fail, operating_time_hours)
    pfd = failure_rates * (8760 * operating_time_years/sum_of_test) / (2 * 1000000)
    return pfd

# Fungsi untuk menentukan SIL berdasarkan PFD Average
def determine_sil(pfd_average):
    if pfd_average >= 0.01:
        return 'SIL-1'
    elif pfd_average < 0.01 and pfd_average >= 0.001:
        return 'SIL-2'
    elif pfd_average < 0.001 and pfd_average >= 0.0001:
        return 'SIL-3'
    elif pfd_average < 0.0001:
        return 'NO FAILURE'

# Fungsi untuk menghitung reliabilitas dan ketersediaan
def calculate_reliability_and_availability(on_demand_fail, operating_time_years, number_of_devices, pfd_average):
    operating_time_hours = operating_time_years * 8760 * number_of_devices
    if operating_time_hours == 0:
        failure_rates_per_hour = 0
    else:
        failure_rates_per_hour = on_demand_fail / operating_time_hours

    mttr = on_demand_fail * 12  # MTTR calculation
    availability = 1 - (failure_rates_per_hour * mttr)
    reliability = 1 - pfd_average

    # Konversi ke format persen dengan membulatkan hingga 4 desimal
    reliability_percent = round(reliability * 100, 4)
    availability_percent = round(availability * 100, 4)

    return reliability_percent, availability_percent

# Deklarasi variabel global untuk menyimpan frame hasil
result_frame = None

def display_sil(sil):
    color = ""
    if sil == 'SIL-1':
        color = "red"
        textcol="white"
    elif sil == 'SIL-2':
        color = "#79AA45"  # Warna kuning untuk SIL-2
        textcol="white"
    elif sil == 'SIL-3':
        color = "lightgreen"
        textcol="white"
    else:
        color = "white"
        textcol="black"

    # Menampilkan SIL dengan warna
    sil_label = customtkinter.CTkLabel(result_frame, text=sil, bg_color=color, font=("Arial Black", 20), text_color=textcol)
    sil_label.pack()

def create_result_frame():
    global result_frame
    if result_frame is not None and result_frame.winfo_exists():
        # Clear the existing frame
        for widget in result_frame.winfo_children():
            widget.destroy()
    else:
        # Create the frame because it doesn't exist
        result_frame = customtkinter.CTkFrame(main_content)
        result_frame.pack(fill='y')

def check_and_display_calculate_button():
    if loaded_file_path and sil_platform_menu.get():
        calculate_sil_button.pack(side='top', pady=10)
    else:
        calculate_sil_button.pack_forget()

def add_log_update_database(platform, package, last_test_date, pass_or_fail):
    log_excel_path = 'log_data_sil_phm.xlsx'
    log_sheet_name = 'Log Update Database'
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not os.path.exists(log_excel_path):
        workbook = Workbook()
        workbook.save(log_excel_path)
    
    workbook = load_workbook(log_excel_path)
    if log_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(log_sheet_name)
    log_sheet = workbook[log_sheet_name]
    
    if log_sheet.max_row == 1:  # Jika ini adalah header atau belum ada data
        # Menulis header dari DataFrame
        log_sheet.append(["Platform", "Package", "Last Test Date", "Update Type", "Nama Pengguna", "Update Timestamp"])
    
    # Menulis data ke dalam worksheet
    log_sheet.append([platform, package, last_test_date.strftime('%Y-%m-%d'), pass_or_fail, user_name, now])
    
    workbook.save(log_excel_path)
    workbook.close()

def log_calculation_to_excel():
    global df_results, user_name
    log_excel_path = 'log_data_sil_phm.xlsx'
    log_sil_calculation_sheet_name = 'Log SIL Calculation'
    
    # Buka workbook atau buat baru jika tidak ada
    if os.path.exists(log_excel_path):
        workbook = load_workbook(log_excel_path)
    else:
        workbook = Workbook()
        workbook.save(log_excel_path)

    # Pastikan sheet untuk Log SIL Calculation ada
    if log_sil_calculation_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(log_sil_calculation_sheet_name)
    worksheet = workbook[log_sil_calculation_sheet_name]

    # Menulis header jika sheet kosong
    if worksheet.max_row == 1:
        headers = list(df_results.columns) + ["Nama Pengguna"]
        worksheet.append(headers)

    # Tambahkan nama pengguna ke baris data
    for row_data in df_results.itertuples(index=False):
        data_with_user = tuple(row_data) + (user_name,)
        worksheet.append(data_with_user)

    workbook.save(log_excel_path)
    workbook.close()

def calculate_and_display_sil_data():
    global df_results, results_treeview
    selected_platform = sil_platform_menu.get()
    if not loaded_file_path or not selected_platform:
        messagebox.showerror("Error", "Silakan unggah file database dan pilih platform.")
        return
    
    create_result_frame()

    try:
        df = pd.read_excel(loaded_file_path)
        required_columns = ['Platform', 'On Demand Fail', 'Operating Years', 'Sum of Test', 'Number of Devices']
        if not all(col in df.columns for col in required_columns):
            messagebox.showerror(
                "Error",
                "File Excel tidak memiliki semua kolom yang dibutuhkan: " + ", ".join(required_columns)
            )
            return
        
        platform_data = df[df['Platform'] == selected_platform]
        total_pfd = 0
        pfd_values = {}
        
        for index, row in platform_data.iterrows():
            on_demand_fail = row['On Demand Fail']
            operating_time_years = row['Operating Years']
            sum_of_test = row['Sum of Test']
            number_of_devices = row['Number of Devices']
            
            # Hitung PFD untuk setiap baris sebelum mencoba menggunakannya
            pfd = calculate_pfd(on_demand_fail, operating_time_years, sum_of_test, number_of_devices)
            
            total_pfd += pfd
            pfd_values[row['Package']] = pfd
        
        pfd_average = total_pfd
        sil = determine_sil(pfd_average)

        # Menampilkan hasil PFD dan SIL di bawah tombol
        for package, pfd_value in pfd_values.items():
            customtkinter.CTkLabel(result_frame, text=f"PFD - {package}: {pfd_value:.2E}", text_color="white", font=("Arial", 16)).pack()
        customtkinter.CTkLabel(result_frame, text=f"PFD Average: {pfd_average:.2E}", text_color="white", font=("Arial", 16, "bold")).pack()

        # Memanggil fungsi display_sil untuk menampilkan SIL dengan warna
        display_sil(sil)
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sites = selected_platform  # Replace with actual logic to get the site
        equipment = selected_platform
        current_package_pm_test_interval = "1Y"  # Example, replace with your logic
        instrument_type = "ESD & PSS LOOP"  # Example, replace with your logic
        current_instr_pm_test_interval = "3M/6M/1Y"  # Example, replace with your logic
        current_pfod = f"{pfd_average:.2E}"  # Example, replace with your logic
        sil_class = sil  # Already defined above
        forecast_sil_class = sil  # Replace with your logic
        compliance = "Compliant" if sil in ["SIL-2", "SIL-3", "NO FAILURE"] else "Non-Compliant"  # Example logic
        proposed_package = "2Y" if sil in ["SIL-2", "SIL-3", "NO FAILURE"] else "1Y"  # Example logic
        reliability, availability = calculate_reliability_and_availability(on_demand_fail, operating_time_years, number_of_devices, pfd_average)
        add_or_update_calculation_result(timestamp, sites, equipment, current_package_pm_test_interval,
                                 instrument_type, current_instr_pm_test_interval,
                                 current_pfod, sil_class, forecast_sil_class, compliance, 
                                 proposed_package, reliability, availability)
        
        # Setelah melakukan perhitungan, pastikan Treeview dan tombol print ditampilkan
        build_and_update_results_treeview()  # Fungsi ini membangun kembali Treeview jika tidak ada atau memperbaruinya

        log_calculation_to_excel()
        
        if not hasattr(main_content, 'print_button_created'):
            display_print_button()

    except Exception as e:
        messagebox.showerror("Error", f"Terjadi kesalahan saat menghitung SIL: {e}")

# Fungsi untuk menampilkan tombol "Print"
def display_print_button():
    # Ciptakan tombol print
    print_button = customtkinter.CTkButton(main_content, text="Print", hover=True, hover_color="orange", command=handle_print)
    print_button.pack(side="bottom", anchor="c")
    # Simpan referensi ke tombol print di main_content agar bisa diakses dari fungsi lain
    main_content.print_button = print_button
    main_content.print_button_created = True

# Di bagian atas, di bawah 'loaded_file_path = None'
calculated_platforms = []  # Menyimpan daftar platform yang telah dihitung

def confirm_sil_calculation():
    global calculated_and_printed, printed
    selected_platform = sil_platform_menu.get()

    # Jika data sudah di-print dan platform belum dipilih lagi, berikan pemberitahuan
    if printed or not selected_platform:
        messagebox.showwarning("Warning", "Please select the platform again before calculating!")
        return

    if not loaded_file_path:
        messagebox.showerror("Error", "Silakan unggah file database terlebih dahulu.")
        return
    
    # Periksa apakah perhitungan sudah dilakukan dan belum di-print
    if calculated_and_printed.get(selected_platform, False):
        messagebox.showinfo("Informasi", "Perhitungan untuk platform ini sudah dilakukan. Silakan print hasil terlebih dahulu atau pilih platform lain.")
        return

    # Munculkan dialog konfirmasi
    confirm = messagebox.askyesno("Konfirmasi", f"Apakah yakin ingin menghitung PFD dan menentukan SIL dari Platform {selected_platform}?")
    if confirm:
        calculate_and_display_sil_data()
        # Setelah perhitungan, tandai sebagai true
        calculated_and_printed[selected_platform] = True
        if not hasattr(main_content, 'print_button_created'):
            display_print_button()  # Menampilkan tombol "Print" jika belum dibuat

def populate_sil_data(*args):
    global calculate_sil_button, printed
    selected_platform = sil_platform_menu.get()

    # Reset printed status ketika platform baru dipilih
    printed = False
    
    # Bersihkan data yang sudah ada sebelum menambahkan yang baru
    for widget in main_content.winfo_children():
        if isinstance(widget, customtkinter.CTkFrame):
            widget.destroy()

    if not selected_platform or not loaded_file_path:
        messagebox.showerror("Error", "Silakan pilih platform terlebih dahulu.")
        return
    
    try:
        df = pd.read_excel(loaded_file_path)
        platform_data = df[df['Platform'] == selected_platform]

        last_widget = None
        for index, row in platform_data.iterrows():
            package_frame = customtkinter.CTkFrame(main_content)
            package_frame.pack(fill='x', padx=10, pady=5)

            customtkinter.CTkLabel(package_frame, text=f"Data for Package: {row['Package']}", text_color="white", font=("Arial", 18, "bold")).pack(anchor='w')
            customtkinter.CTkLabel(package_frame, text=f"Operating Time (Years): {row['Operating Years']}", text_color="white", font=("Arial", 16)).pack(anchor='w')
            customtkinter.CTkLabel(package_frame, text=f"Number of Devices (Population): {row['Number of Devices']}", text_color="white", font=("Arial", 16)).pack(anchor='w')
            customtkinter.CTkLabel(package_frame, text=f"On Demand (Test): {row.get('Sum of Test', 'N/A')}", text_color="white", font=("Arial", 16)).pack(anchor='w')
            customtkinter.CTkLabel(package_frame, text=f"On Demand (Fail): {row.get('On Demand Fail', 'N/A')}", text_color="white", font=("Arial", 16)).pack(anchor='w')
            last_widget = package_frame
        # Tempatkan tombol "Calculate PFDs and Determine SIL" setelah data terakhir
        if last_widget is not None:
            calculate_sil_button.pack(side='top', pady=10, after=last_widget)
    except Exception as e:
        messagebox.showerror("Error", f"Terjadi kesalahan saat membaca file Excel: {e}")

template_file_path = None

# Function to determine the base directory of the application
def get_base_path():
    if getattr(sys, 'frozen', False):
        # If the application is run as a bundled executable, the sys._MEIPASS
        # attribute contains the path to the temporary directory that PyInstaller
        # uses to store the bundled application.
        return sys._MEIPASS
    else:
        # Otherwise, just use the current directory
        return os.path.dirname(os.path.abspath(__file__))
    
def save_to_excel():
    global df_results, template_file_path, calculated_and_printed, printed, user_name
    try:
        # Dapatkan timestamp saat ini dan format menjadi string
        current_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        # Tentukan nama file baru dengan timestamp
        new_excel_path = f"report_data_{current_timestamp}.xlsx"
        
        # Copy template.xlsx ke new_excel_path yang sudah diformat
        shutil.copyfile(template_file_path, new_excel_path)
        
        # Load workbook yang baru saja disalin
        workbook = load_workbook(filename=new_excel_path)
        # Pilih sheet kedua dengan index [1] untuk ditulis
        worksheet = workbook.worksheets[1]

        # Menulis header dari DataFrame
        for col_num, column_title in enumerate(df_results.columns, start=1):
            worksheet.cell(row=1, column=col_num).value = column_title
        
        # Menambahkan header untuk nama pengguna
        worksheet.cell(row=1, column=len(df_results.columns) + 1).value = "Nama Pengguna"
        
        # Menulis data ke dalam worksheet mulai dari baris kedua
        for row_num, row_data in enumerate(df_results.itertuples(index=False), start=2):
            for col_num, cell_value in enumerate(row_data, start=1):
                worksheet.cell(row=row_num, column=col_num).value = cell_value
            # Menambahkan nama pengguna ke setiap baris
            worksheet.cell(row=row_num, column=len(df_results.columns) + 1).value = user_name
        
        # Simpan workbook yang telah diupdate
        workbook.save(new_excel_path)
        workbook.close()

        # Reset DataFrame setelah disimpan
        df_results = pd.DataFrame(columns=[
            "Timestamp", "Sites", "Equipment", "Current Package PM Test Interval",
            "Instrument Type", "Current Instr PM Test Interval",
            "Current Probability Failure on Demand Average",
            "SIL Class - Current ESD1 Test Interval (Baseline 1Y)",
            "Forecast SIL Class at 2Y ESD1 Test Interval",
            "Compliance of SIL at forecast 2Y ESD1 test interval",
            "Proposed Package", "Reliability", "Availability"
        ])
        
        # Update results_treeview untuk menunjukkan bahwa DataFrame sudah direset
        update_results_treeview()

        # Tampilkan pesan sukses
        messagebox.showinfo("Success", f"Data has been successfully exported to {new_excel_path}.")

        printed = True

        # Sembunyikan Treeview dan tombol print
        results_treeview.pack_forget()
        if hasattr(main_content, 'print_button_created'):
            main_content.print_button.pack_forget()
            delattr(main_content, 'print_button_created')

        # Setelah berhasil menyimpan ke Excel, reset status perhitungan dan pencetakan
        for platform in calculated_and_printed.keys():
            calculated_and_printed[platform] = False
        
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

# Perbarui fungsi print_confirmation untuk memanggil fungsi save_to_excel
def print_confirmation():
    # Buat window baru untuk konfirmasi
    confirm_window = customtkinter.CTkToplevel()
    confirm_window.title("Print Data")

    # Metode ini membuat jendela konfirmasi menjadi transient untuk root window
    confirm_window.transient(root)

    # Ini akan mengarahkan semua event ke jendela konfirmasi
    confirm_window.grab_set()

    # Ini akan memaksa jendela konfirmasi untuk mengambil fokus
    confirm_window.focus_force()
    
    # Tambahkan pesan konfirmasi
    customtkinter.CTkLabel(confirm_window, text="Apakah Anda yakin ingin mencetak dan mengekspor data?").pack(pady=10)

    # Tambahkan tombol konfirmasi dan batal
    customtkinter.CTkButton(confirm_window, text="Ya, Saya Yakin", command=lambda: [save_to_excel(), confirm_window.destroy()]).pack(side='left', padx=(20, 10), pady=10)
    customtkinter.CTkButton(confirm_window, text="Tidak, Batalkan", command=confirm_window.destroy).pack(side='right', padx=(10, 20), pady=10)

    # Metode ini akan menunggu sampai jendela konfirmasi ditutup sebelum melanjutkan eksekusi
    confirm_window.wait_window()

# Fungsi untuk menangani pencetakan
def handle_print():
    # Tampilkan konfirmasi
    print_confirmation()

if __name__ == "__main__":
    base_path = get_base_path()
    template_file_path = os.path.join(base_path, 'template_sil_report.xlsx')
    customtkinter.set_appearance_mode("dark")
    root = customtkinter.CTk()
    root.title("Safety Integrity Level Calculation Dashboard")
    root.geometry("768x384")

    # Sidebar setup
    sidebar = customtkinter.CTkFrame(root, width=500)
    sidebar.pack(side="left", fill="both", padx=5, pady=5)

    # Main content setup
    main_content = customtkinter.CTkScrollableFrame(master=root, scrollbar_button_hover_color="red")
    main_content.pack(side="left", fill="both", expand=True, padx=5, pady=5)  # Change fill to 'both'

    # Tempatkan kode ini di tempat yang sesuai di dalam if __name__ == "__main__":
    calculate_sil_button = customtkinter.CTkButton(main_content, text="Calculate PFDs and Determine SIL",
                                     command=confirm_sil_calculation, hover=True, hover_color="orange")

    # Add file upload button and label in the sidebar
    upload_label = customtkinter.CTkLabel(sidebar, text="Upload Database (Excel)", font=("Arial", 16, "bold"))
    upload_label.pack(fill='x')
    upload_button = customtkinter.CTkButton(sidebar, text="Browse files", command=lambda: upload_file(), hover=True, text_color="white")
    upload_button.pack(fill='x', pady=5)

    upload_success_label = customtkinter.CTkLabel(sidebar, text="", font=("Arial", 16))
    if not loaded_file_path:
        upload_success_label.pack_forget()

    # New label for "UPDATE DATABASE STROKING TEST"
    update_label = customtkinter.CTkLabel(sidebar, text="UPDATE DATABASE", font=("Arial", 16, "bold"))

    # Dropdown for platform in the "UPDATE DATABASE STROKING TEST" section
    platform_var = tk.StringVar()
    platform_label = customtkinter.CTkLabel(sidebar, text="Select Platform", font=("Arial", 16))
    platform_menu = customtkinter.CTkComboBox(sidebar, state="normal", variable=platform_var, button_hover_color="#d93636", hover=True)

    package_var = tk.StringVar()
    package_label = customtkinter.CTkLabel(sidebar, text="Select Package", font=("Arial", 16))
    package_menu = customtkinter.CTkComboBox(sidebar, variable=package_var, state='normal', button_hover_color="#d93636", hover=True)

    # Last test date entry with mindate set to December 31, 2022
    last_test_label = customtkinter.CTkLabel(sidebar, text="Last Test Date", font=("Arial", 16))
    last_test_var = tk.StringVar()
    last_test_entry = DateEntry(sidebar, textvariable=last_test_var, date_pattern='yyyy-mm-dd', mindate=datetime(2022, 12, 31),
                                font=("Arial", 20))

    # Pass and Fail buttons
    pass_button = customtkinter.CTkButton(sidebar, text="Pass", command=lambda: pass_test(), hover=True, hover_color="#32CD32")
    # Modifikasi tombol Fail untuk memanggil fungsi show_on_demand_fail_input
    fail_button = customtkinter.CTkButton(sidebar, text="Fail", command=fail_test, hover=True, hover_color="#FF4500")

    # Add a label and entry for the On Demand Fail input, but don't pack them yet
    fail_input_label = customtkinter.CTkLabel(sidebar, text="On Demand Fail", font=("Arial", 16))
    fail_input_var = tk.IntVar()
    fail_input = ttk.Spinbox(sidebar, from_=0, to=100, textvariable=fail_input_var, wrap=True, font=("Arial", 20))

    # Add a Confirm button for the On Demand Fail input, but don't pack it yet
    confirm_fail_button = customtkinter.CTkButton(sidebar, text="Confirm Fail Update", command=confirm_fail_update)

    # New label for "INPUT DATA FOR SIL CALCULATION"
    sil_calculation_label = customtkinter.CTkLabel(sidebar, text="INPUT DATA FOR SIL CALCULATION", font=("Arial", 16, "bold"))

    # Dropdown for platform in the "INPUT DATA FOR SIL CALCULATION" section
    sil_platform_label = customtkinter.CTkLabel(sidebar, text="Select Platform for SIL", font=("Arial", 16))
    sil_platform_menu = customtkinter.CTkComboBox(sidebar, state='normal', hover=True, button_hover_color="red", command=populate_sil_data)

    show_initial_dialog()

    root.mainloop()